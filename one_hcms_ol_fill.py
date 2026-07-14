"""
One HCMS, Offering Letter fill engine.
Maps an approved OL payload to the official KSNI 2026 Excel template, routed by
grade tier and gender. Label based cell lookup, so it tolerates row shifts
between tiers. Keeps Total Remunerasi and THR as template formulas, and keeps
all formatting. Produces a single sheet final letter.

Usage:
  from one_hcms_ol_fill import fill_ol
  fill_ol("2026_Offering_Template_KSNI_revA.xlsx", data, "OL_final.xlsx")

data = {
  "grade": "5A", "gender": "L" | "P",
  "nomorSeq": "01471", "signDate": "2026-08-01",
  "nama": "...", "alamat": "...",
  "position": "...", "division": "...", "golongan": "5A", "location": "Bandung",
  "components": { "Gaji Pokok": 15000000, "Tunjangan Jabatan": 5000000, ... }
}
"""
import openpyxl

ROMAN = ["", "I", "II", "III", "IV", "V", "VI", "VII", "VIII", "IX", "X", "XI", "XII"]
BULAN = ["", "Januari", "Februari", "Maret", "April", "Mei", "Juni", "Juli",
         "Agustus", "September", "Oktober", "November", "Desember"]


def ol_tier(grade):
    try:
        n = int("".join(ch for ch in str(grade) if ch.isdigit())[:1])
    except Exception:
        n = 0
    if n >= 5:
        return "G5UP"
    if n == 4:
        return "G4"
    if n == 3:
        return "G3"
    return "LOW"


def ol_sheet(grade, gender):
    t = ol_tier(grade)
    if t == "G5UP":
        return "Standard Grade 5 Up"
    if t == "G4":
        return "Standard Grade 4 (L)" if gender == "L" else "Standard Grade 4 (P)"
    return "Standard Grade 3"


def gen_nomor(seq, date_iso):
    y, m, d = [int(x) for x in date_iso.split("-")]
    return str(seq) + "/KSNI-HRD/OFF/" + ROMAN[m] + "/" + str(y)


def long_date(date_iso):
    y, m, d = [int(x) for x in date_iso.split("-")]
    return str(d) + " " + BULAN[m] + " " + str(y)


def _rows_with(ws, col, text, exact=False):
    hits = []
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=col).value
        if v is None:
            continue
        s = str(v).strip()
        if (s == text) if exact else (s.startswith(text)):
            hits.append(r)
    return hits


def _set_offer(ws, label_col_b, value):
    # offer fields, label in col B, value in col E
    hits = _rows_with(ws, 2, label_col_b)
    if hits:
        ws.cell(row=hits[0], column=5, value=value)


def from_ol_payload(ol, nomor_seq, sign_date):
    """Build the fill data straight from an approved OL payload.
    Identity comes from the frozen snapshot, so nothing is retyped."""
    idn = ol.get("identity") or {}
    bd = ol.get("basicData") or {}
    tpl = ol.get("olTemplate") or {}
    props = ol.get("proposals") or {}
    final_key = props.get("final") or "p1"
    final = props.get(final_key) or {}
    labels = {"gapok": "Gaji Pokok", "tjab": "Tunjangan Jabatan", "thadir": "Tunjangan Kehadiran",
              "tkom": "Tunjangan Komunikasi", "ttrans": "Tunjangan Transportasi", "tkend": "Tunjangan Pengganti Kendaraan"}
    comps = {}
    for k, v in (final.get("components") or {}).items():
        if k in labels and v:
            comps[labels[k]] = v
    customs = final.get("customs") or {}
    extras = []
    for c in (tpl.get("customComponents") or []):
        amt = customs.get(c.get("key")) or 0
        if amt:
            extras.append({"name": c.get("name"), "amount": amt, "nature": c.get("nature"), "showOnLetter": c.get("showOnLetter")})
    return {
        "grade": ol.get("grade"), "gender": idn.get("gender") or bd.get("gender") or "L",
        "nomorSeq": nomor_seq, "signDate": sign_date,
        "nama": idn.get("legalName", ""), "alamat": idn.get("address", ""),
        "position": ol.get("finalPosition", ""), "division": bd.get("division", ""),
        "golongan": ol.get("grade"), "location": bd.get("workLocation", ""),
        "status": bd.get("statusKekaryawanan"),
        "components": comps, "customComponents": extras
    }


def fill_ol(template_path, data, out_path):
    wb = openpyxl.load_workbook(template_path)
    sheet = ol_sheet(data.get("grade"), data.get("gender"))
    ws = wb[sheet]

    # 1. Nomor, cell in col A containing KSNI-HRD
    nomor = data.get("nomor") or gen_nomor(data.get("nomorSeq", "00000"), data.get("signDate", "2026-01-01"))
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if v and "KSNI-HRD" in str(v):
            ws.cell(row=r, column=1, value=nomor)
            break

    # 2. Identity, Pihak Kedua. Nama appears twice in col A, take the last.
    nama_rows = _rows_with(ws, 1, "Nama", exact=True)
    if nama_rows:
        ws.cell(row=nama_rows[-1], column=5, value=data.get("nama", ""))
    alamat_rows = _rows_with(ws, 1, "Alamat", exact=True)
    if alamat_rows:
        ws.cell(row=alamat_rows[0], column=5, value=data.get("alamat", ""))

    # 3. Offer fields, label col B, value col E
    _set_offer(ws, "Posisi Jabatan", data.get("position", ""))
    _set_offer(ws, "Divisi", data.get("division", ""))
    _set_offer(ws, "Golongan", data.get("golongan", data.get("grade", "")))
    _set_offer(ws, "Lokasi Kerja", data.get("location", "Bandung"))
    if data.get("status"):
        _set_offer(ws, "Status Kekaryawanan", data.get("status"))

    # 4. Components, label col C, besaran col F. Keep Total and THR formulas.
    comps = data.get("components", {})
    for label, amount in comps.items():
        for r in range(1, ws.max_row + 1):
            v = ws.cell(row=r, column=3).value
            if v and str(v).strip().startswith(label):
                ws.cell(row=r, column=6, value=amount)
                break

    # 4b. Komponen tambahan penambah yang bertanda tampil di surat.
    # Disisipkan sebagai baris sebelum Total Remunerasi, gaya sel disalin dari
    # baris komponen terakhir agar sama persis, lalu rumus Total ditulis ulang.
    # THR tetap dari tiga komponen pertama, F31 sampai F33, tidak terpengaruh.
    extras = [c for c in (data.get("customComponents") or []) if c.get("nature") == "ADD" and c.get("showOnLetter") and float(c.get("amount") or 0) > 0]
    if extras:
        from copy import copy
        trows = _rows_with(ws, 2, "Total Remunerasi")
        if trows:
            total_row = trows[0]
            for ex in extras:
                ws.insert_rows(total_row)
                src = total_row - 1
                for m in list(ws.merged_cells.ranges):
                    if str(m) == "B" + str(total_row) + ":E" + str(total_row):
                        ws.merged_cells.ranges.discard(m)
                for col in range(1, 8):
                    s = ws.cell(row=src, column=col)
                    dcell = ws.cell(row=total_row, column=col)
                    if s.has_style:
                        dcell.font = copy(s.font); dcell.border = copy(s.border); dcell.fill = copy(s.fill)
                        dcell.number_format = s.number_format; dcell.alignment = copy(s.alignment); dcell.protection = copy(s.protection)
                prev_no = ws.cell(row=src, column=2).value
                try:
                    no = int(prev_no) + 1
                except Exception:
                    no = ""
                ws.cell(row=total_row, column=2, value=str(no))
                ws.cell(row=total_row, column=3, value=ex.get("name"))
                ws.cell(row=total_row, column=6, value=float(ex.get("amount")))
                total_row += 1
            ws.cell(row=total_row, column=6, value="=SUM(F31:F" + str(total_row - 1) + ")")

    # 5. Sign date, cell in col G containing Bandung,
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=7).value
        if v and str(v).strip().startswith("Bandung,"):
            ws.cell(row=r, column=7, value="Bandung,  " + long_date(data.get("signDate", "2026-01-01")))
            break

    # 6. Single sheet final letter, drop the other tiers
    for other in list(wb.sheetnames):
        if other != sheet:
            del wb[other]

    wb.save(out_path)
    return {"sheet": sheet, "nomor": nomor, "out": out_path}
